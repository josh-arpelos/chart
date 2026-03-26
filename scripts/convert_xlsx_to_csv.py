"""Convert source xlsx files to CSV metadata and results files."""

import csv
import os
import re

import openpyxl
import pandas as pd

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_DIR = os.path.join(ROOT, "data", "source_xlsx")
METADATA_DIR = os.path.join(ROOT, "metadata")
RESULTS_DIR = os.path.join(ROOT, "data", "results")
RAW_DIR = os.path.join(ROOT, "data", "raw")

os.makedirs(METADATA_DIR, exist_ok=True)
os.makedirs(RESULTS_DIR, exist_ok=True)


def parse_platemap():
    """Parse platemap xlsx into a well_id -> target mapping.

    The platemap has identical layouts for all 4 donors (rows 3-10).
    8 rows (A-H) x 8 columns (1-8) = 64 wells.
    """
    wb = openpyxl.load_workbook(os.path.join(SOURCE_DIR, "platemap.xlsx"))
    ws = wb.active

    # Extract from first donor block (rows 3-10: A through H, columns B-I for cols 1-8)
    platemap = []
    for row_idx in range(3, 11):  # rows 3-10
        row_letter = ws.cell(row=row_idx, column=1).value  # A-H
        for col_idx in range(2, 10):  # columns B-I (plate columns 1-8)
            target = ws.cell(row=row_idx, column=col_idx).value
            if target is None:
                continue
            plate_col = col_idx - 1
            well_id = f"{row_letter}{plate_col}"
            is_isotype = "isotype" in target.lower()
            platemap.append({
                "well_id": well_id,
                "row": row_letter,
                "column": plate_col,
                "target": target,
                "is_isotype_control": is_isotype,
            })

    # Also parse ICS/Treg block (rows 47-50, columns 1-4)
    treg_platemap = []
    for row_idx in range(47, 51):
        row_letter = ws.cell(row=row_idx, column=1).value
        for col_idx in range(2, 6):  # columns B-E (plate columns 1-4)
            target = ws.cell(row=row_idx, column=col_idx).value
            if target is None:
                continue
            plate_col = col_idx - 1
            well_id = f"{row_letter}{plate_col}"
            treg_platemap.append({
                "well_id": well_id,
                "row": row_letter,
                "column": plate_col,
                "target": target,
                "is_isotype_control": False,
            })

    return platemap, treg_platemap


def write_platemap_csv(platemap, treg_platemap):
    """Write metadata/platemap.csv."""
    path = os.path.join(METADATA_DIR, "platemap.csv")
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["well_id", "row", "column", "target", "is_isotype_control"])
        writer.writeheader()
        for entry in platemap:
            writer.writerow(entry)
    print(f"Wrote {len(platemap)} wells to {path}")

    # Write Treg platemap separately
    treg_path = os.path.join(METADATA_DIR, "treg_platemap.csv")
    with open(treg_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["well_id", "row", "column", "target", "is_isotype_control"])
        writer.writeheader()
        for entry in treg_platemap:
            writer.writerow(entry)
    print(f"Wrote {len(treg_platemap)} wells to {treg_path}")


def convert_compiled_data():
    """Convert compiled data xlsx to CSV with snake_case columns."""
    path = os.path.join(SOURCE_DIR, "compiled_data_vFinal.xlsx")
    df = pd.read_excel(path, sheet_name="Sheet1")

    column_map = {
        "Target": "target",
        "Target Classification": "target_classification",
        "Well ID": "well_id",
        "Donor ID": "donor_id",
        "B Cells % Receptor Expression": "b_cells_pct_expression",
        "NK Cells % Receptor Expression": "nk_cells_pct_expression",
        "CD4+ Cells % Receptor Expression": "cd4_cells_pct_expression",
        "CD8+ Cells % Receptor Expression": "cd8_cells_pct_expression",
        "B Cells MFI": "b_cells_mfi",
        "NK Cells MFI": "nk_cells_mfi",
        "CD4+ Cells MFI": "cd4_cells_mfi",
        "CD8+ Cells MFI": "cd8_cells_mfi",
        "B Cells Receptor Density": "b_cells_receptor_density",
        "NK Cells Receptor Density": "nk_cells_receptor_density",
        "CD4+ Cells Receptor Density": "cd4_cells_receptor_density",
        "CD8+ Cells Receptor Density": "cd8_cells_receptor_density",
    }

    df = df.rename(columns=column_map)
    df = df.dropna(how="all")
    df["donor_id"] = df["donor_id"].astype(int)

    out_path = os.path.join(RESULTS_DIR, "compiled_data.csv")
    df.to_csv(out_path, index=False)
    print(f"Wrote {len(df)} rows to {out_path}")
    return df


def write_targets_csv(compiled_df, platemap):
    """Write metadata/targets.csv from compiled data + platemap."""
    # Build well_id -> target mapping from platemap
    well_to_target = {e["well_id"]: e for e in platemap}

    targets = compiled_df[["target", "target_classification"]].drop_duplicates().sort_values("target")
    targets["is_isotype_control"] = targets["target"].str.contains("isotype", case=False)

    path = os.path.join(METADATA_DIR, "targets.csv")
    targets.to_csv(path, index=False)
    print(f"Wrote {len(targets)} targets to {path}")


def write_donors_csv():
    """Write metadata/donors.csv."""
    donors = []
    for donor_dir in sorted(os.listdir(RAW_DIR)):
        donor_path = os.path.join(RAW_DIR, donor_dir)
        if not os.path.isdir(donor_path):
            continue
        fcs_files = [f for f in os.listdir(donor_path) if f.endswith(".fcs")]
        donors.append({
            "donor_id": donor_dir,
            "n_fcs_files": len(fcs_files),
        })

    path = os.path.join(METADATA_DIR, "donors.csv")
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["donor_id", "n_fcs_files"])
        writer.writeheader()
        for d in donors:
            writer.writerow(d)
    print(f"Wrote {len(donors)} donors to {path}")


def write_fcs_manifest(platemap, treg_platemap):
    """Write metadata/fcs_manifest.csv by walking data/raw/."""
    well_to_target = {e["well_id"]: e["target"] for e in platemap}
    treg_well_to_target = {e["well_id"]: e["target"] for e in treg_platemap}

    manifest = []
    for donor_dir in sorted(os.listdir(RAW_DIR)):
        donor_path = os.path.join(RAW_DIR, donor_dir)
        if not os.path.isdir(donor_path):
            continue
        for filename in sorted(os.listdir(donor_path)):
            if not filename.endswith(".fcs"):
                continue

            rel_path = f"data/raw/{donor_dir}/{filename}"

            # Parse file type and well from filename
            if "Beads_stained" in filename:
                file_type = "beads_stained"
                well_id = ""
                target = ""
            elif "Beads_us" in filename:
                file_type = "beads_unstained"
                well_id = ""
                target = ""
            else:
                file_type = "sample"
                # Extract well from end of filename: ..._A1.fcs
                match = re.search(r"_([A-H]\d+)\.fcs$", filename)
                well_id = match.group(1) if match else ""
                if donor_dir == "Treg":
                    target = treg_well_to_target.get(well_id, "")
                else:
                    target = well_to_target.get(well_id, "")

            manifest.append({
                "filename": filename,
                "donor_id": donor_dir,
                "well_id": well_id,
                "file_type": file_type,
                "relative_path": rel_path,
                "target": target,
            })

    path = os.path.join(METADATA_DIR, "fcs_manifest.csv")
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["filename", "donor_id", "well_id", "file_type", "relative_path", "target"])
        writer.writeheader()
        for entry in manifest:
            writer.writerow(entry)
    print(f"Wrote {len(manifest)} FCS entries to {path}")


if __name__ == "__main__":
    platemap, treg_platemap = parse_platemap()
    write_platemap_csv(platemap, treg_platemap)
    compiled_df = convert_compiled_data()
    write_targets_csv(compiled_df, platemap)
    write_donors_csv()
    write_fcs_manifest(platemap, treg_platemap)
    print("\nDone!")
